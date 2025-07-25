import pandas as pd
from utils.connpd import execute_query
from utils.connpp import execute_queryPP
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import requests
from bs4 import BeautifulSoup

import openpyxl
from openpyxl import load_workbook


select_columns = """
    TABLE_NAME
    ,TABLE_SCHEMA+'.'+TABLE_NAME as TABLE_FULL_NAME
	,COLUMN_NAME
	,DATA_TYPE	
 	,IS_NULLABLE
	,TABLE_CATALOG 
    ,GetDate() as Date
    ,ORDINAL_POSITION
	"""
#####################
###################
## Connect to the Data Lake and extract the catalog of JDE_BI_OPS
DL_biops_DataLake = execute_query(f"""
SELECT 
	{select_columns}
        
FROM [RDL00001_EnterpriseDataLanding].INFORMATION_SCHEMA.COLUMNS
WHERE TABLE_SCHEMA = 'JDE_BI_OPS' AND TABLE_NAME NOT LIKE '%Test%'
ORDER BY TABLE_NAME,COLUMN_NAME""")


# connect to shema dbo Datawarehouse
DW_dboPD = execute_query(f"""
SELECT 
	{select_columns}
FROM [RDL00001_EnterpriseDataWarehouse].INFORMATION_SCHEMA.COLUMNS
WHERE TABLE_SCHEMA = 'dbo' AND TABLE_NAME NOT LIKE '%Test%'
ORDER BY TABLE_NAME,COLUMN_NAME""")


DW_dboPP = execute_queryPP(f"""
SELECT 
	{select_columns}
FROM [RDL00001_EnterpriseDataWarehouse].INFORMATION_SCHEMA.COLUMNS
WHERE TABLE_SCHEMA = 'dbo' AND TABLE_NAME NOT LIKE '%Test%' 
ORDER BY TABLE_NAME,COLUMN_NAME""")


# connect to shema dbo Biops2 en PP
# Cataloge_Biops2 = execute_queryPP(f"""
# SELECT
# 	{select_columns}
# FROM [RDL00001_EnterpriseDataWarehouse].INFORMATION_SCHEMA.COLUMNS
# WHERE TABLE_SCHEMA = 'dbo' AND TABLE_NAME NOT LIKE '%Test%' """)


# connect to shema JDE DataLanding_jde
DataLanding_jde = execute_query(f"""
SELECT 
	{select_columns}
FROM 
    [RDL00001_EnterpriseDataLanding].INFORMATION_SCHEMA.COLUMNS
WHERE 
    TABLE_SCHEMA = 'JDE' 
    AND TABLE_NAME NOT LIKE '%Test%'
    AND TABLE_NAME NOT LIKE '%copy%'
    AND TABLE_NAME NOT LIKE '%bkp%'
    ORDER BY TABLE_NAME,COLUMN_NAME
    """)
###########################GET THE EXCEL REPORT AND CONFIGURED IT########################
Excel_file = "catalog.xlsx"
# create / replace the file########
# with pd.ExcelWriter(
#     Excel_file, engine="openpyxl", mode="w"  # Use "w" to overwrite
# ) as writer:
#     DL_biops_DataLake.to_excel(writer, sheet_name="DL_biops_DataLake", startrow=2, index=False)
#     DW_dboPP.to_excel(writer, sheet_name="DW_dboPP", startrow=2, index=False)
#     DW_dboPD.to_excel(writer, sheet_name="DW_dboPD", startrow=2, index=False)
#     DataLanding_jde.to_excel(writer, sheet_name="DataLanding_jde", startrow=2, index=False)
###########


# Write multiple DataFrames to separate sheets in an existing Excel file using openpyxl
with pd.ExcelWriter(
    Excel_file, engine="openpyxl", mode="a", if_sheet_exists="replace"
) as writer:
    DL_biops_DataLake.to_excel(
        writer, sheet_name="DL_biops_DataLake", startrow=2, index=False
    )

    DW_dboPP.to_excel(writer, sheet_name="DW_dboPP", startrow=2, index=False)

    DW_dboPD.to_excel(writer, sheet_name="DW_dboPD", startrow=2, index=False)

    DataLanding_jde.to_excel(
        writer, sheet_name="DataLanding_jde", startrow=2, index=False
    )

# Load the updated Excel workbook to modify titles and column widths
wb = load_workbook(Excel_file)

# Dictionary mapping sheet names to their descriptive titles (used in cell A1 of each sheet)
titles = {
    "DL_biops_DataLake": "🧬 Views _BIOPS RDL00001_EnterpriseDataLanding / schema: JDE_BI_OPS",
    "DW_dboPP": "🏥 Tables en PP dans RDL00001_EnterpriseDataWarehouse  / schema: dbo",
    "DW_dboPD": "🧾 Tables en PD  dans RDL00001_EnterpriseDataWarehouse  / schema: dbo",
    "DataLanding_jde": "Tables dans RDL00001_EnterpriseDataLanding   /  schema: JDE",
}

# Set horizontal alignment for cells (currently set to left)
alignment = Alignment(horizontal="left")  # Options include "left", "center", "right"

# Loop over each sheet and title
for sheet_name, title in titles.items():
    ws = wb[sheet_name]  # Access the worksheet by name
    ws["A1"] = title  # Insert title into cell A1

    # Adjust column widths based on max length of values, skipping title rows (row 1 and 2)
    for col_cells in ws.iter_cols(min_row=3):
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)

        for cell in col_cells:
            try:
                cell_value = str(cell.value)
                if cell_value:
                    max_length = max(max_length, len(cell_value))  # Track max length
            except:
                pass  # Safely ignore errors while accessing cell values

        # Set column width with a bit of extra padding
        adjusted_width = max_length + 2
        ws.column_dimensions[col_letter].width = adjusted_width

# Save the changes to the workbook
wb.save(Excel_file)

############################################################################################
